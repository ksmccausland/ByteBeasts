import os
import random
import re
import sys
from pathlib import Path
from typing import List, Tuple, Set

import tkinter as tk
from tkinter import messagebox

from PIL import Image, ImageTk
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter


# -----------------------------
# Configuration
# -----------------------------
IMAGE_FOLDER = Path("lidar_imgs")          # Folder containing ~20 images
NUM_IMAGES_TO_SHOW = 5                      # Show exactly 5 unique images
GRID_ROWS = 4
GRID_COLS = 4
WINDOW_TITLE = "LIDAR Image Selector"
INSTRUCTION_TEXT = "Click the part of the image with squares in it.\n(You can select multiple grid cells; click again to unselect.)"
EXCEL_FILENAME = Path("lidar_images.xlsx")  # Output Excel file

# Canvas sizing (image will be scaled to fit while preserving aspect ratio)
CANVAS_MAX_W = 900
CANVAS_MAX_H = 700

# Selected cell border appearance
SELECTED_BORDER_COLOR = "#1E90FF"  # DodgerBlue
SELECTED_BORDER_WIDTH = 3
UNSELECTED_BORDER_COLOR = "#CCCCCC"
UNSELECTED_BORDER_WIDTH = 1


# -----------------------------
# Utility functions
# -----------------------------
def find_images(folder: Path) -> List[Path]:
    exts = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".gif"}
    return [p for p in folder.iterdir() if p.suffix.lower() in exts and p.is_file()]


def extract_image_number(fname: str) -> int:
    """
    Extract the first integer found in the filename.
    If none found, return -1 (we'll still record filename as fallback).
    """
    m = re.search(r"(\d+)", fname)
    return int(m.group(1)) if m else -1


def ensure_excel_headers(wb_path: Path):
    """
    Ensure the workbook exists with the proper headers:
    Column A: image_number
    Column B: filename
    Columns C..: cell_1 .. cell_16
    """
    headers = ["image_number", "filename"]
    for i in range(1, GRID_ROWS * GRID_COLS + 1):
        headers.append(f"cell_{i}")

    if wb_path.exists():
        # Verify headers exist (create if new empty file somehow)
        try:
            wb = load_workbook(wb_path)
            ws = wb.active
            if ws.max_row == 0 or ws.max_column == 0:
                ws.append(headers)
                wb.save(wb_path)
            else:
                # If headers missing/wrong length, we won’t rewrite; we’ll append in current shape
                pass
            wb.close()
        except Exception:
            # If file unreadable, back it up and recreate
            backup = wb_path.with_suffix(".backup.xlsx")
            try:
                os.replace(wb_path, backup)
            except Exception:
                pass
            wb = Workbook()
            ws = wb.active
            ws.title = "results"
            ws.append(headers)
            wb.save(wb_path)
            wb.close()
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "results"
        ws.append(headers)
        wb.save(wb_path)
        wb.close()


def append_result_to_excel(wb_path: Path, image_path: Path, selected_indices: Set[int]):
    """
    Append a row:
    image_number | filename | cell_1..cell_16 (1 or 0)
    Indices are 0..15 left->right, top->bottom
    """
    ensure_excel_headers(wb_path)

    image_num = extract_image_number(image_path.name)
    row = [image_num, image_path.name]
    total_cells = GRID_ROWS * GRID_COLS
    for i in range(total_cells):
        row.append(1 if i in selected_indices else 0)

    wb = load_workbook(wb_path)
    ws = wb.active
    ws.append(row)

    # Optional: autosize columns a bit (lightweight)
    for col in range(1, len(row) + 1):
        letter = get_column_letter(col)
        # Set a reasonable width; not perfect autosize but helps
        ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 10, 14)

    wb.save(wb_path)
    wb.close()


# -----------------------------
# GUI Application
# -----------------------------
class ImageGridSelector(tk.Tk):
    def __init__(self, image_paths: List[Path]):
        super().__init__()
        self.title(WINDOW_TITLE)

        # State
        self.image_paths = image_paths
        self.index = 0  # which image we’re on
        self.selected_cells: Set[int] = set()  # currently selected cell indices for the displayed image

        # UI
        self.instruction = tk.Label(self, text=INSTRUCTION_TEXT, font=("Arial", 12))
        self.instruction.pack(pady=(10, 6))

        # Canvas for image + grid
        self.canvas = tk.Canvas(self, width=CANVAS_MAX_W, height=CANVAS_MAX_H, bg="black", highlightthickness=0)
        self.canvas.pack(padx=10, pady=8)

        # Buttons
        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=(2, 12))

        self.submit_btn = tk.Button(btn_frame, text="Submit Selection", command=self.on_submit, width=20)
        self.submit_btn.pack(side=tk.LEFT, padx=8)

        self.quit_btn = tk.Button(btn_frame, text="Quit", command=self.on_quit, width=10)
        self.quit_btn.pack(side=tk.LEFT, padx=8)

        # Bind clicks
        self.canvas.bind("<Button-1>", self.on_canvas_click)

        # Loaded image and geometry
        self.original_img = None         # PIL Image
        self.tk_img = None               # ImageTk
        self.display_w = None
        self.display_h = None
        self.offset_x = 0
        self.offset_y = 0

        # Start with first image
        self.show_current_image()

    # -------- Image loading and drawing --------
    def show_current_image(self):
        self.selected_cells.clear()
        img_path = self.image_paths[self.index]

        # Load and scale image to fit canvas while preserving aspect ratio
        pil_img = Image.open(img_path).convert("RGB")
        self.original_img = pil_img

        can_w, can_h = CANVAS_MAX_W, CANVAS_MAX_H
        img_w, img_h = pil_img.size

        scale = min(can_w / img_w, can_h / img_h)
        disp_w = int(img_w * scale)
        disp_h = int(img_h * scale)

        self.display_w, self.display_h = disp_w, disp_h
        self.offset_x = (can_w - disp_w) // 2
        self.offset_y = (can_h - disp_h) // 2

        resized = pil_img.resize((disp_w, disp_h), Image.BILINEAR)
        self.tk_img = ImageTk.PhotoImage(resized)

        # Clear canvas and draw image
        self.canvas.delete("all")
        self.canvas.create_image(self.offset_x, self.offset_y, anchor=tk.NW, image=self.tk_img, tags="img")

        # Draw grid
        self.draw_grid()

        # Window title update
        self.title(f"{WINDOW_TITLE}  —  {img_path.name}  ({self.index+1}/{len(self.image_paths)})")

    def draw_grid(self):
        # Base light borders for all cells
        cell_w = self.display_w / GRID_COLS
        cell_h = self.display_h / GRID_ROWS

        for r in range(GRID_ROWS):
            for c in range(GRID_COLS):
                i = r * GRID_COLS + c
                x1 = self.offset_x + c * cell_w
                y1 = self.offset_y + r * cell_h
                x2 = x1 + cell_w
                y2 = y1 + cell_h
                # Draw base unselected border
                self.canvas.create_rectangle(
                    x1, y1, x2, y2,
                    outline=UNSELECTED_BORDER_COLOR,
                    width=UNSELECTED_BORDER_WIDTH,
                    tags=f"cell_{i}"
                )

        # Draw selection overlays (blue) on top
        self.redraw_selected_borders()

    def redraw_selected_borders(self):
        # Remove old selection borders
        self.canvas.delete("sel_border")

        cell_w = self.display_w / GRID_COLS
        cell_h = self.display_h / GRID_ROWS

        for i in self.selected_cells:
            r = i // GRID_COLS
            c = i % GRID_COLS
            x1 = self.offset_x + c * cell_w
            y1 = self.offset_y + r * cell_h
            x2 = x1 + cell_w
            y2 = y1 + cell_h

            # Blue border for selected
            self.canvas.create_rectangle(
                x1, y1, x2, y2,
                outline=SELECTED_BORDER_COLOR,
                width=SELECTED_BORDER_WIDTH,
                tags="sel_border"
            )

    # -------- Interaction handlers --------
    def on_canvas_click(self, event):
        # Check if click is within image area
        if not (self.offset_x <= event.x <= self.offset_x + self.display_w and
                self.offset_y <= event.y <= self.offset_y + self.display_h):
            return

        # Map click to grid cell
        rel_x = event.x - self.offset_x
        rel_y = event.y - self.offset_y

        cell_w = self.display_w / GRID_COLS
        cell_h = self.display_h / GRID_ROWS

        c = int(rel_x // cell_w)
        r = int(rel_y // cell_h)

        # Safety clamp
        c = max(0, min(GRID_COLS - 1, c))
        r = max(0, min(GRID_ROWS - 1, r))

        idx = r * GRID_COLS + c

        # Toggle selection
        if idx in self.selected_cells:
            self.selected_cells.remove(idx)
        else:
            self.selected_cells.add(idx)

        self.redraw_selected_borders()

    def on_submit(self):
        # Save the current image's selections to Excel
        img_path = self.image_paths[self.index]
        append_result_to_excel(EXCEL_FILENAME, img_path, self.selected_cells)

        # Advance to next image or finish
        if self.index + 1 < len(self.image_paths):
            self.index += 1
            self.show_current_image()
        else:
            messagebox.showinfo("Done", f"Selections saved to '{EXCEL_FILENAME.name}'. The program will now exit.")
            self.destroy()

    def on_quit(self):
        if messagebox.askokcancel("Quit", "Are you sure you want to quit? Unsaved selections for this image will be lost."):
            self.destroy()


# -----------------------------
# Main
# -----------------------------
def main():
    # Validate images
    if not IMAGE_FOLDER.exists():
        print(f"Error: Folder '{IMAGE_FOLDER}' not found.", file=sys.stderr)
        sys.exit(1)

    all_images = find_images(IMAGE_FOLDER)
    if len(all_images) == 0:
        print(f"Error: No images found in '{IMAGE_FOLDER}'.", file=sys.stderr)
        sys.exit(1)

    # Choose 5 unique images (or as many as possible up to 5)
    k = min(NUM_IMAGES_TO_SHOW, len(all_images))
    image_choice = random.sample(all_images, k)

    # Ensure Excel has headers ready
    ensure_excel_headers(EXCEL_FILENAME)

    # Launch app
    app = ImageGridSelector(image_choice)
    app.mainloop()


if __name__ == "__main__":
    main()
